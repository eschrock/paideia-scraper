from openpyxl import Workbook
from openpyxl.styles import Font


class Output:
    """Handles output operations for the Paideia scraper."""

    def __init__(self):
        """Initialize the Output class."""
        pass

    def _optimize_column_widths(
        self, worksheet, headers, students_in_class, class_name
    ):
        """
        Optimize column widths based on content length and provide reasonable defaults.

        Args:
            worksheet: The worksheet to optimize
            headers: List of column headers
            students_in_class: List of student data for this class
            class_name: Name of the class (for logging)
        """
        # Define reasonable minimum and maximum widths for different column types
        min_widths = {
            "Student": 15,
            "Class": 20,
            "Name": 20,  # Parent names
            "Email": 30,  # Email addresses
            "Phone": 15,  # Phone numbers
        }

        max_widths = {"Student": 40, "Class": 30, "Name": 40, "Email": 50, "Phone": 20}

        # Calculate optimal widths for each column
        for col_idx, header in enumerate(headers, 1):
            # Determine column type for width constraints
            if "Student" in header:
                col_type = "Student"
            elif "Class" in header:
                col_type = "Class"
            elif "Name" in header:
                col_type = "Name"
            elif "Email" in header:
                col_type = "Email"
            elif "Phone" in header:
                col_type = "Phone"
            else:
                col_type = "Name"  # Default fallback

            # Get minimum and maximum widths for this column type
            min_width = min_widths.get(col_type, 15)
            max_width = max_widths.get(col_type, 30)

            # Calculate optimal width based on content
            optimal_width = len(header)  # Start with header length

            # Check all data in this column
            # Skip header row
            for row_idx in range(2, len(students_in_class) + 2):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # Add some padding and account for the actual content length
                    content_length = len(str(cell_value)) + 2
                    optimal_width = max(optimal_width, content_length)

            # Apply constraints
            final_width = max(min_width, min(optimal_width, max_width))

            # Set the column width
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = final_width

    def write_excel(self, students, excel_path):
        """
        Write student data to an Excel file with separate sheets for each class.
        Parents are collapsed into columns (Parent 1 Name, Parent 1 Email, etc.).

        Args:
            students: List of student dictionaries with name, class, and parents
            excel_path: Path to the Excel file to create/overwrite
        """
        # Create a new workbook
        workbook = Workbook()

        # Remove the default sheet
        workbook.remove(workbook.active)

        # Group students by class and find max parents per class
        class_data = {}
        max_parents_per_class = {}

        for student in students:
            class_name = student["class"]
            student_name = student["name"]
            parent_count = len(student["parents"])

            # Initialize class data if it doesn't exist
            if class_name not in class_data:
                class_data[class_name] = []
                max_parents_per_class[class_name] = 0

            # Update max parents for this class
            max_parents_per_class[class_name] = max(
                max_parents_per_class[class_name], parent_count
            )

            # Add student data
            class_data[class_name].append(
                {"name": student_name, "parents": student["parents"]}
            )

        # Create a sheet for each class
        for class_name, students_in_class in class_data.items():
            max_parents = max_parents_per_class[class_name]

            # Create worksheet
            worksheet = workbook.create_sheet(title=class_name)

            # Generate headers dynamically
            headers = ["Student", "Class"]
            for i in range(1, max_parents + 1):
                headers.extend(
                    [f"Parent {i} Name", f"Parent {i} Email", f"Parent {i} Phone"]
                )

            # Add header row
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            # Add data rows (one row per student)
            for row_idx, student_data in enumerate(students_in_class, 2):
                # Student name and class
                worksheet.cell(row=row_idx, column=1, value=student_data["name"])
                worksheet.cell(row=row_idx, column=2, value=class_name)

                # Parent information
                for parent_idx, parent in enumerate(student_data["parents"], 1):
                    # Skip Student and Class columns
                    parent_col_start = (parent_idx - 1) * 3
                    col_offset = 2 + parent_col_start

                    worksheet.cell(
                        row=row_idx, column=col_offset + 1, value=parent["name"]
                    )
                    worksheet.cell(
                        row=row_idx, column=col_offset + 2, value=parent["email"]
                    )
                    worksheet.cell(
                        row=row_idx, column=col_offset + 3, value=parent["phone"]
                    )

                # Fill remaining parent columns with None for students with fewer parents
                for parent_idx in range(
                    len(student_data["parents"]) + 1, max_parents + 1
                ):
                    parent_col_start = (parent_idx - 1) * 3
                    col_offset = 2 + parent_col_start

                    worksheet.cell(row=row_idx, column=col_offset + 1, value=None)
                    worksheet.cell(row=row_idx, column=col_offset + 2, value=None)
                    worksheet.cell(row=row_idx, column=col_offset + 3, value=None)

            # Optimize column widths based on content
            self._optimize_column_widths(
                worksheet, headers, students_in_class, class_name
            )

        # Save the workbook
        workbook.save(excel_path)
